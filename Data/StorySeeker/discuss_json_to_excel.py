import json
import ast
import pandas as pd
import argparse
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Parse command line arguments
parser = argparse.ArgumentParser(
    description='Convert discuss JSON file to Excel format')
parser.add_argument('-i', '--input', type=str,
                    default=r'F:\Work\Debate\PrivousPerspectiveCoder-LM\Data\StorySeeker\gpt-4o_output\discuss_process\json\baseline2.json',
                    help='Input JSON file path')
parser.add_argument('-o', '--output', type=str,
                    default=r'F:\Work\Debate\PrivousPerspectiveCoder-LM\Data\StorySeeker\gpt-4o_output\discuss_process\excel\baseline2.xlsx',
                    help='Output Excel file path')
args = parser.parse_args()

# Read JSON file
input_file = args.input
output_file = args.output

print(f"读取输入文件: {input_file}")
with open(input_file, 'r', encoding='utf-8') as f:
    data = json.load(f)

# Create Excel writer object
writer = pd.ExcelWriter(output_file, engine='openpyxl')

# 1. Role_Team sheet - Display role info and codes in separate rows
role_team_rows = []
for idx, team_member in enumerate(data['Role_Team']):
    # First code item: includes complete role information
    if team_member['init_codebook']:
        first_code = team_member['init_codebook'][0]
        first_row = {
            'Gender': team_member.get('Gender', ''),
            'Education': team_member.get('Education', ''),
            'Race_Ethnicity': team_member.get('Race_Ethnicity', ''),
            'Age': team_member.get('Age', ''),
            'Degree_Subject': team_member.get('Degree_Subject', ''),
            'Reddit_Use': team_member.get('Reddit_Use', ''),
            'Positionality': team_member.get('positionality', ''),
            'Code': first_code['code'],
            'Definition': first_code['definition'],
            'Inclusion Criteria': '\n'.join(first_code['inclusion_criteria']),
            'Exclusion Criteria': '\n'.join(first_code['exclusion_criteria']),
            'Typical Examples': '\n'.join(first_code['typical_examples']),
            'Atypical Examples': '\n'.join(first_code['atypical_examples']),
            'Participants': ', '.join(first_code['participants']),
            'Relevance to RQ': first_code['relevance_to_RQ'],
            'Notes': first_code['notes']
        }
        role_team_rows.append(first_row)

        # Remaining code items: leave role info columns empty
        for code_item in team_member['init_codebook'][1:]:
            code_row = {
                'Gender': '',
                'Education': '',
                'Race_Ethnicity': '',
                'Age': '',
                'Degree_Subject': '',
                'Reddit_Use': '',
                'Positionality': '',
                'Code': code_item['code'],
                'Definition': code_item['definition'],
                'Inclusion Criteria': '\n'.join(code_item['inclusion_criteria']),
                'Exclusion Criteria': '\n'.join(code_item['exclusion_criteria']),
                'Typical Examples': '\n'.join(code_item['typical_examples']),
                'Atypical Examples': '\n'.join(code_item['atypical_examples']),
                'Participants': ', '.join(code_item['participants']),
                'Relevance to RQ': code_item['relevance_to_RQ'],
                'Notes': code_item['notes']
            }
            role_team_rows.append(code_row)

df_role_team = pd.DataFrame(role_team_rows)
df_role_team.to_excel(writer, sheet_name='Role_Team', index=False)

# 2. Agreed_disagreed_codebook sheet - Agreement and Disagreement each occupy one row
agreed_disagreed_rows = []

# Add Agreement main row and sub-rows
if data['Agreed_disagreed_codebook']['agreements']:
    # First agreement includes Type information
    first_agreement = data['Agreed_disagreed_codebook']['agreements'][0]
    first_row = {
        'Type': 'Agreement',
        'Code': first_agreement['code'],
        'Roles': ', '.join(first_agreement['roles']),
        'Reason': first_agreement['reason']
    }
    agreed_disagreed_rows.append(first_row)

    # Remaining agreements as sub-rows
    for agreement in data['Agreed_disagreed_codebook']['agreements'][1:]:
        row = {
            'Type': '',
            'Code': agreement['code'],
            'Roles': ', '.join(agreement['roles']),
            'Reason': agreement['reason']
        }
        agreed_disagreed_rows.append(row)

# Add Disagreement main row and sub-rows
if data['Agreed_disagreed_codebook']['disagreements']:
    # First disagreement includes Type information
    first_disagreement = data['Agreed_disagreed_codebook']['disagreements'][0]
    first_row = {
        'Type': 'Disagreement',
        'Code': first_disagreement['code'],
        'Roles': ', '.join(first_disagreement['roles']),
        'Reason': first_disagreement['reason']
    }
    agreed_disagreed_rows.append(first_row)

    # Remaining disagreements as sub-rows
    for disagreement in data['Agreed_disagreed_codebook']['disagreements'][1:]:
        row = {
            'Type': '',
            'Code': disagreement['code'],
            'Roles': ', '.join(disagreement['roles']),
            'Reason': disagreement['reason']
        }
        agreed_disagreed_rows.append(row)

df_agreed_disagreed = pd.DataFrame(agreed_disagreed_rows)
df_agreed_disagreed.to_excel(writer, sheet_name='Agreed_disagreed_codebook', index=False)

# 3. Discussion rounds - one sheet per round
discussion_data = data.get('Discussion', {})

def parse_round(round_value):
    if isinstance(round_value, str):
        try:
            return ast.literal_eval(round_value)
        except (ValueError, SyntaxError):
            return []
    if isinstance(round_value, list):
        return round_value
    return []

rounds_payload = {
    'round1': parse_round(discussion_data.get('round1')),
    'round2': parse_round(discussion_data.get('round2')),
    'round3': parse_round(discussion_data.get('round3'))
}

for round_name, round_list in rounds_payload.items():
    round_rows = []
    if round_name == 'round1':
        for role_block in round_list:
            if not isinstance(role_block, dict):
                continue
            for _, role_payload in role_block.items():
                role_name = role_payload.get('Role', '')
                evidences = role_payload.get('Disagreement_Evidence', [])
                for ev in evidences:
                    lit = ev.get('Evidence_Literature', {})
                    content = ev.get('Evidence_Content', {})
                    logic = ev.get('Evidence_Logic', {})
                    round_rows.append({
                        'Role': role_name,
                        'Code': ev.get('code', ''),
                        'Preliminary_Claim': ev.get('Preliminary_Claim', ''),
                        'Definition': ev.get('Definition', ''),
                        'Evidence_Literature/Author': lit.get('Author', ''),
                        'Evidence_Literature/Year': lit.get('Year', ''),
                        'Evidence_Literature/Title': lit.get('Title', ''),
                        'Evidence_Literature/DOI_or_Link': lit.get('DOI_or_Link', ''),
                        'Evidence_Literature/Summary': lit.get('Summary', ''),
                        'Evidence_Content/Observation': content.get('Observation', ''),
                        'Evidence_Content/Relevance': content.get('Relevance', ''),
                        'Evidence_Logic/Reasoning': logic.get('Reasoning', ''),
                        'Evidence_Logic/Impact': logic.get('Impact', '')
                    })
    elif round_name == 'round2':
        for role_block in round_list:
            if not isinstance(role_block, dict):
                continue
            for _, role_payload in role_block.items():
                role_name = role_payload.get('Role', '')
                decision_pattern = role_payload.get('DecisionPattern', '')
                reasoning_list = role_payload.get('Reasoning', [])
                for item in reasoning_list:
                    self_check = item.get('Self_Check', {})
                    cross = item.get('Cross_Discussion', {})
                    round_rows.append({
                        'Role': role_name,
                        'DecisionPattern': decision_pattern,
                        'Code': item.get('code', ''),
                        'Disagreement_Point': item.get('Disagreement_Point', ''),
                        'Self_Check/First_ExplicitStance': self_check.get('First_ExplicitStance', ''),
                        'Self_Check/Then_DefinitionConsistency': self_check.get('Then_DefinitionConsistency', ''),
                        'Self_Check/Finally_LimitationsDecision': self_check.get('Finally_LimitationsDecision', ''),
                        'Cross_Discussion/PerRoleAnalysis': json.dumps(cross.get('PerRoleAnalysis', []), ensure_ascii=False),
                        'Cross_Discussion/Consensus': cross.get('Consensus', ''),
                        'Cross_Discussion/Divergence': cross.get('Divergence', ''),
                        'Cross_Discussion/HybridPerspective': cross.get('HybridPerspective', '')
                    })
    elif round_name == 'round3':
        for role_block in round_list:
            if not isinstance(role_block, dict):
                continue
            for _, role_payload in role_block.items():
                role_name = role_payload.get('Role', '')
                final_results = role_payload.get('Final_Results', [])
                for result in final_results:
                    self_ref = result.get('Self_Reflection', {})
                    round_rows.append({
                        'Role': role_name,
                        'Code': result.get('code', ''),
                        'Initial_Claim_Round1': result.get('Initial_Claim_Round1', ''),
                        'Final_Decision': result.get('Final_Decision', ''),
                        'Final_Claim': result.get('Final_Claim', ''),
                        'Final_Explanation': result.get('Final_Explanation', ''),
                        'Self_Reflection/Strengths': self_ref.get('Strengths', ''),
                        'Self_Reflection/Weaknesses': self_ref.get('Weaknesses', ''),
                        'Self_Reflection/Future_Consideration': self_ref.get('Future_Consideration', '')
                    })
    df_round = pd.DataFrame(round_rows)
    df_round.to_excel(writer, sheet_name=round_name, index=False)

# 4. Final_codebook sheet
final_codebook_rows = []
for code_item in data['Final_codebook']['codebook']:
    row = {
        'Code': code_item['code'],
        'Definition': code_item['definition'],
        'Inclusion Criteria': '\n'.join(code_item['inclusion_criteria']),
        'Exclusion Criteria': '\n'.join(code_item['exclusion_criteria']),
        'Typical Examples': '\n'.join(code_item['typical_examples']),
        'Atypical Examples': '\n'.join(code_item['atypical_examples']),
        'Participants': ', '.join(code_item['participants']),
        'Relevance to RQ': code_item['relevance_to_RQ'],
        'Notes': code_item['notes']
    }
    final_codebook_rows.append(row)

df_final = pd.DataFrame(final_codebook_rows)
df_final.to_excel(writer, sheet_name='Final_codebook', index=False)

writer.close()

# 格式化Excel
wb = load_workbook(output_file)

# Define border style
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Set format for each sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    header_rows = 1

    # Set header row style
    for row_idx in range(1, header_rows + 1):
        for cell in ws[row_idx]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    # Special handling for Role_Team sheet - merge role info cells
    if sheet_name == 'Role_Team':
        current_row = header_rows + 1
        while current_row <= ws.max_row:
            role_cell = ws.cell(row=current_row, column=1)

            if role_cell.value:  # Main row with role information
                # Calculate how many codes this role has (number of sub-rows)
                merge_count = 1
                check_row = current_row + 1
                while check_row <= ws.max_row:
                    check_cell = ws.cell(row=check_row, column=1)
                    if not check_cell.value:  # Empty row indicates sub-row
                        merge_count += 1
                        check_row += 1
                    else:  # Encountered next role
                        break

                # Merge role info columns (first 7 columns)
                if merge_count > 1:
                    for col_idx in range(1, 8):  # Gender to Positionality columns
                        ws.merge_cells(start_row=current_row, start_column=col_idx,
                                       end_row=current_row + merge_count - 1, end_column=col_idx)

                # Set all cell styles (white background, black border, vertical center)
                for row_offset in range(merge_count):
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=current_row + row_offset, column=col_idx)
                        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
                        cell.border = thin_border

                current_row += merge_count
            else:
                current_row += 1
    elif sheet_name == 'Agreed_disagreed_codebook':
        # Special handling for Agreed_disagreed_codebook sheet - merge Type column
        current_row = header_rows + 1
        while current_row <= ws.max_row:
            type_cell = ws.cell(row=current_row, column=1)

            if type_cell.value:  # Main row with Type information (Agreement or Disagreement)
                # Calculate how many codes this type has (number of sub-rows)
                merge_count = 1
                check_row = current_row + 1
                while check_row <= ws.max_row:
                    check_cell = ws.cell(row=check_row, column=1)
                    if not check_cell.value:  # Empty row indicates sub-row
                        merge_count += 1
                        check_row += 1
                    else:  # Encountered next type
                        break

                # Merge Type column (1st column)
                if merge_count > 1:
                    ws.merge_cells(start_row=current_row, start_column=1,
                                   end_row=current_row + merge_count - 1, end_column=1)

                # Set all cell styles
                for row_offset in range(merge_count):
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=current_row + row_offset, column=col_idx)
                        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
                        cell.border = thin_border

                current_row += merge_count
            else:
                current_row += 1
    elif sheet_name in {'round1', 'round2', 'round3'}:
        # Merge Role column for round sheets
        current_row = header_rows + 1
        while current_row <= ws.max_row:
            role_cell = ws.cell(row=current_row, column=1)

            if role_cell.value:
                merge_count = 1
                check_row = current_row + 1
                while check_row <= ws.max_row:
                    check_cell = ws.cell(row=check_row, column=1)
                    if check_cell.value == role_cell.value:
                        merge_count += 1
                        check_row += 1
                    else:
                        break

                if merge_count > 1:
                    ws.merge_cells(start_row=current_row, start_column=1,
                                   end_row=current_row + merge_count - 1, end_column=1)

                for row_offset in range(merge_count):
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=current_row + row_offset, column=col_idx)
                        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
                        cell.border = thin_border

                current_row += merge_count
            else:
                current_row += 1
    else:
        # Regular format for other sheets
        for row in ws.iter_rows(min_row=header_rows + 1):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
                cell.border = thin_border

    # Auto-adjust column width
    for idx, col in enumerate(ws.columns, 1):
        column = get_column_letter(idx)
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        # Set wider column width to show more content
        if sheet_name == 'Role_Team':
            if idx == 7:  # Positionality column
                adjusted_width = max(80, min(max_length + 2, 140))
            elif idx >= 8:  # Code and subsequent columns
                adjusted_width = max(80, min(max_length + 2, 140))
            else:
                adjusted_width = max(25, min(max_length + 2, 80))
        else:
            adjusted_width = max(25, min(max_length + 2, 140))

        ws.column_dimensions[column].width = adjusted_width

    # Set row height to fit content
    if sheet_name == 'Role_Team':
        # For Role_Team, set uniform row height for each code row (non-merged cell area)
        for row_idx in range(header_rows + 1, ws.max_row + 1):
            # Only calculate content from code columns (7th column and beyond)
            max_lines = 1
            for col_idx in range(8, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    text = str(cell.value)
                    lines = text.count('\n') + 1
                    col_width = ws.column_dimensions[get_column_letter(col_idx)].width
                    estimated_lines = max(1, int(len(text) / (col_width)))
                    total_lines = lines + estimated_lines
                    max_lines = max(max_lines, total_lines)

            # Set uniform row height
            ws.row_dimensions[row_idx].height = max(30, min(max_lines * 18, 800))
    elif sheet_name == 'Agreed_disagreed_codebook':
        # For Agreed_disagreed_codebook, only calculate non-Type column content
        for row_idx in range(header_rows + 1, ws.max_row + 1):
            max_lines = 1
            for col_idx in range(2, ws.max_column + 1):  # Start from 2nd column
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    text = str(cell.value)
                    lines = text.count('\n') + 1
                    col_width = ws.column_dimensions[get_column_letter(col_idx)].width
                    estimated_lines = max(1, int(len(text) / (col_width)))
                    total_lines = lines + estimated_lines
                    max_lines = max(max_lines, total_lines)

            ws.row_dimensions[row_idx].height = max(30, min(max_lines * 18, 800))
    else:
        # Normal row height calculation for other sheets
        for row_idx in range(header_rows + 1, ws.max_row + 1):
            max_lines = 1
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    text = str(cell.value)
                    lines = text.count('\n') + 1
                    col_width = ws.column_dimensions[get_column_letter(col_idx)].width
                    estimated_lines = max(1, int(len(text) / (col_width * 1.2)))
                    total_lines = lines + estimated_lines
                    max_lines = max(max_lines, total_lines)

            ws.row_dimensions[row_idx].height = max(30, min(max_lines * 18, 800))

    # Freeze first row
    ws.freeze_panes = f'A{header_rows + 1}'

wb.save(output_file)
print(f"\n转换完成！Excel文件已保存为: {output_file}")
print(f"输入文件: {input_file}")
print(f"输出文件: {output_file}")
print(f"\n创建的工作表:")
print("1. Role_Team - 角色团队的分层代码本（角色信息为主行，代码为子行）")
print("2. Agreed_disagreed_codebook - 一致与不一致的代码")
print("3. round1 - 讨论第1轮（分表）")
print("4. round2 - 讨论第2轮（分表）")
print("5. round3 - 讨论第3轮（分表）")
print("6. Final_codebook - 最终代码本")
