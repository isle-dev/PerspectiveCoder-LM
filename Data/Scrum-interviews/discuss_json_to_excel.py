import json
import pandas as pd
import argparse
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Parse command line arguments
parser = argparse.ArgumentParser(
    description='Convert discuss JSON file to Excel format (supports hierarchical structure)')
parser.add_argument('-i', '--input', type=str,
                    default='F:\Work\Debate\PrivousPerspectiveCoder-LM\Data\Scrum-interviews\gpt-4o_output\discuss_process\json\discuss_same_per.json',
                    help='Input JSON file path')
parser.add_argument('-o', '--output', type=str,
                    default='F:\Work\Debate\PrivousPerspectiveCoder-LM\Data\Scrum-interviews\gpt-4o_output\discuss_process\excel\discuss_same_per.xlsx',
                    help='Output Excel file path')
args = parser.parse_args()

# Read JSON file
input_file = args.input
output_file = args.output

print(f"读取输入文件: {input_file}")
with open(input_file, 'r', encoding='utf-8') as f:
    data = json.load(f)

# Create Excel writer object
writer = pd.ExcelWriter(output_file, engine='openpyxl')

# 1. Role_Team sheet - Display role info and codes in separate rows
role_team_rows = []
for idx, team_member in enumerate(data['Role_Team']):
    # First code item: includes complete role information
    if team_member['init_codebook']:
        first_code = team_member['init_codebook'][0]
        first_row = {
            'Role': team_member['role'],
            'Study Level': team_member['Intended_Study_Level'],
            'Subject': team_member['Subject'],
            'Research Interest': team_member['Research_Interest'],
            'Dimensions Source': team_member['Dimensions_Source'],
            'Positionality': team_member['positionality'],
            'Code': first_code['code'],
            'Definition': first_code['definition'],
            'Inclusion Criteria': '\n'.join(first_code['inclusion_criteria']),
            'Exclusion Criteria': '\n'.join(first_code['exclusion_criteria']),
            'Typical Examples': '\n'.join(first_code['typical_examples']),
            'Atypical Examples': '\n'.join(first_code['atypical_examples']),
            'Participants': ', '.join(first_code['participants']),
            'Relevance to RQ': first_code['relevance_to_RQ'],
            'Notes': first_code['notes']
        }
        role_team_rows.append(first_row)

        # Remaining code items: leave role info columns empty
        for code_item in team_member['init_codebook'][1:]:
            code_row = {
                'Role': '',
                'Study Level': '',
                'Subject': '',
                'Research Interest': '',
                'Dimensions Source': '',
                'Positionality': '',
                'Code': code_item['code'],
                'Definition': code_item['definition'],
                'Inclusion Criteria': '\n'.join(code_item['inclusion_criteria']),
                'Exclusion Criteria': '\n'.join(code_item['exclusion_criteria']),
                'Typical Examples': '\n'.join(code_item['typical_examples']),
                'Atypical Examples': '\n'.join(code_item['atypical_examples']),
                'Participants': ', '.join(code_item['participants']),
                'Relevance to RQ': code_item['relevance_to_RQ'],
                'Notes': code_item['notes']
            }
            role_team_rows.append(code_row)

df_role_team = pd.DataFrame(role_team_rows)
df_role_team.to_excel(writer, sheet_name='Role_Team', index=False)

# 2. Agreed_disagreed_codebook sheet - Agreement and Disagreement each occupy one row
agreed_disagreed_rows = []

# Add Agreement main row and sub-rows
if data['Agreed_disagreed_codebook']['agreements']:
    # First agreement includes Type information
    first_agreement = data['Agreed_disagreed_codebook']['agreements'][0]
    first_row = {
        'Type': 'Agreement',
        'Code': first_agreement['code'],
        'Roles': ', '.join(first_agreement['roles']),
        'Reason': first_agreement['reason']
    }
    agreed_disagreed_rows.append(first_row)

    # Remaining agreements as sub-rows
    for agreement in data['Agreed_disagreed_codebook']['agreements'][1:]:
        row = {
            'Type': '',
            'Code': agreement['code'],
            'Roles': ', '.join(agreement['roles']),
            'Reason': agreement['reason']
        }
        agreed_disagreed_rows.append(row)

# Add Disagreement main row and sub-rows
if data['Agreed_disagreed_codebook']['disagreements']:
    # First disagreement includes Type information
    first_disagreement = data['Agreed_disagreed_codebook']['disagreements'][0]
    first_row = {
        'Type': 'Disagreement',
        'Code': first_disagreement['code'],
        'Roles': ', '.join(first_disagreement['roles']),
        'Reason': first_disagreement['reason']
    }
    agreed_disagreed_rows.append(first_row)

    # Remaining disagreements as sub-rows
    for disagreement in data['Agreed_disagreed_codebook']['disagreements'][1:]:
        row = {
            'Type': '',
            'Code': disagreement['code'],
            'Roles': ', '.join(disagreement['roles']),
            'Reason': disagreement['reason']
        }
        agreed_disagreed_rows.append(row)

df_agreed_disagreed = pd.DataFrame(agreed_disagreed_rows)
df_agreed_disagreed.to_excel(writer, sheet_name='Agreed_disagreed_codebook', index=False)

# 3. Decision_agreed_codebook sheet - Discussion Result and Decision Agreement each occupy one row
decision_rows = []

# Add Discussion Result main row and sub-rows
if 'discussion_results' in data.get('Decision_agreed_codebook', {}):
    results = data['Decision_agreed_codebook']['discussion_results']
    if results:
        # First result includes Type information
        first_result = results[0]
        first_row = {
            'Type': 'Discussion Result',
            'Code': first_result['code'],
            'Data Source{Literature Evidence}': first_result['data_source'].get('literature_evidence', ''),
            'Data Source{Content Evidence}': first_result['data_source'].get('content_evidence', ''),
            'Data Source{Logic Evidence}': first_result['data_source'].get('logic_evidence', ''),
            'Reasoning': first_result.get('reasoning', ''),
            'Resolution': first_result.get('resolution', ''),
            'Justification': first_result.get('justification', '')
        }
        decision_rows.append(first_row)

        # Remaining results as sub-rows
        for result in results[1:]:
            row = {
                'Type': '',
                'Code': result['code'],
                'Data Source{Literature Evidence}': result['data_source'].get('literature_evidence', ''),
                'Data Source{Content Evidence}': result['data_source'].get('content_evidence', ''),
                'Data Source{Logic Evidence}': result['data_source'].get('logic_evidence', ''),
                'Reasoning': result.get('reasoning', ''),
                'Resolution': result.get('resolution', ''),
                'Justification': result.get('justification', '')
            }
            decision_rows.append(row)

# Add Decision Agreement main row and sub-rows
if 'decision_agreement_codebook' in data.get('Decision_agreed_codebook', {}):
    agreements = data['Decision_agreed_codebook']['decision_agreement_codebook']
    if agreements:
        # First agreement includes Type information
        first_agreement = agreements[0]
        first_row = {
            'Type': 'Decision Agreement',
            'Code': first_agreement['code'],
            'Literature Evidence': '',
            'Content Evidence': '',
            'Logic Evidence': '',
            'Reasoning': '',
            'Resolution': '',
            'Justification': ''
        }
        decision_rows.append(first_row)

        # Remaining agreements as sub-rows
        for code_item in agreements[1:]:
            row = {
                'Type': '',
                'Code': code_item['code'],
                'Literature Evidence': '',
                'Content Evidence': '',
                'Logic Evidence': '',
                'Reasoning': '',
                'Resolution': '',
                'Justification': ''
            }
            decision_rows.append(row)

df_decision = pd.DataFrame(decision_rows)
df_decision.to_excel(writer, sheet_name='Decision_agreed_codebook', index=False)

# 4. Final_codebook sheet
final_codebook_rows = []
for code_item in data['Final_codebook']['codebook']:
    row = {
        'Code': code_item['code'],
        'Definition': code_item['definition'],
        'Inclusion Criteria': '\n'.join(code_item['inclusion_criteria']),
        'Exclusion Criteria': '\n'.join(code_item['exclusion_criteria']),
        'Typical Examples': '\n'.join(code_item['typical_examples']),
        'Atypical Examples': '\n'.join(code_item['atypical_examples']),
        'Participants': ', '.join(code_item['participants']),
        'Relevance to RQ': code_item['relevance_to_RQ'],
        'Notes': code_item['notes']
    }
    final_codebook_rows.append(row)

df_final = pd.DataFrame(final_codebook_rows)
df_final.to_excel(writer, sheet_name='Final_codebook', index=False)

writer.close()

# 格式化Excel
wb = load_workbook(output_file)

# Define border style
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Set format for each sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # Set header row style
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Special handling for Role_Team sheet - merge role info cells
    if sheet_name == 'Role_Team':
        current_row = 2
        while current_row <= ws.max_row:
            role_cell = ws.cell(row=current_row, column=1)

            if role_cell.value:  # Main row with role information
                # Calculate how many codes this role has (number of sub-rows)
                merge_count = 1
                check_row = current_row + 1
                while check_row <= ws.max_row:
                    check_cell = ws.cell(row=check_row, column=1)
                    if not check_cell.value:  # Empty row indicates sub-row
                        merge_count += 1
                        check_row += 1
                    else:  # Encountered next role
                        break

                # Merge role info columns (first 6 columns)
                if merge_count > 1:
                    for col_idx in range(1, 7):  # Role to Positionality columns
                        ws.merge_cells(start_row=current_row, start_column=col_idx,
                                       end_row=current_row + merge_count - 1, end_column=col_idx)

                # Set all cell styles (white background, black border, vertical center)
                for row_offset in range(merge_count):
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=current_row + row_offset, column=col_idx)
                        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
                        cell.border = thin_border

                current_row += merge_count
            else:
                current_row += 1
    elif sheet_name == 'Agreed_disagreed_codebook':
        # Special handling for Agreed_disagreed_codebook sheet - merge Type column
        current_row = 2
        while current_row <= ws.max_row:
            type_cell = ws.cell(row=current_row, column=1)

            if type_cell.value:  # Main row with Type information (Agreement or Disagreement)
                # Calculate how many codes this type has (number of sub-rows)
                merge_count = 1
                check_row = current_row + 1
                while check_row <= ws.max_row:
                    check_cell = ws.cell(row=check_row, column=1)
                    if not check_cell.value:  # Empty row indicates sub-row
                        merge_count += 1
                        check_row += 1
                    else:  # Encountered next type
                        break

                # Merge Type column (1st column)
                if merge_count > 1:
                    ws.merge_cells(start_row=current_row, start_column=1,
                                   end_row=current_row + merge_count - 1, end_column=1)

                # Set all cell styles
                for row_offset in range(merge_count):
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=current_row + row_offset, column=col_idx)
                        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
                        cell.border = thin_border

                current_row += merge_count
            else:
                current_row += 1
    elif sheet_name == 'Decision_agreed_codebook':
        # Special handling for Decision_agreed_codebook sheet - merge Type column
        current_row = 2
        while current_row <= ws.max_row:
            type_cell = ws.cell(row=current_row, column=1)

            if type_cell.value:  # Main row with Type information
                # Calculate how many codes this type has (number of sub-rows)
                merge_count = 1
                check_row = current_row + 1
                while check_row <= ws.max_row:
                    check_cell = ws.cell(row=check_row, column=1)
                    if not check_cell.value:  # Empty row indicates sub-row
                        merge_count += 1
                        check_row += 1
                    else:  # Encountered next type
                        break

                # Merge Type column (1st column)
                if merge_count > 1:
                    ws.merge_cells(start_row=current_row, start_column=1,
                                   end_row=current_row + merge_count - 1, end_column=1)

                # Set all cell styles
                for row_offset in range(merge_count):
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=current_row + row_offset, column=col_idx)
                        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
                        cell.border = thin_border

                current_row += merge_count
            else:
                current_row += 1
    else:
        # Regular format for other sheets
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
                cell.border = thin_border

    # Auto-adjust column width
    for idx, col in enumerate(ws.columns, 1):
        column = get_column_letter(idx)
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        # Set reasonable column width
        if sheet_name == 'Role_Team':
            if idx == 6:  # Positionality column
                adjusted_width = 50
            elif idx >= 7:  # Code and subsequent columns
                adjusted_width = 40
            else:
                adjusted_width = min(max_length + 2, 20)
        else:
            adjusted_width = min(max_length + 2, 60)

        ws.column_dimensions[column].width = adjusted_width

    # Set row height to fit content
    if sheet_name == 'Role_Team':
        # For Role_Team, set uniform row height for each code row (non-merged cell area)
        for row_idx in range(2, ws.max_row + 1):
            # Only calculate content from code columns (7th column and beyond)
            max_lines = 1
            for col_idx in range(7, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    text = str(cell.value)
                    lines = text.count('\n') + 1
                    col_width = ws.column_dimensions[get_column_letter(col_idx)].width
                    estimated_lines = max(1, int(len(text) / (col_width)))
                    total_lines = lines + estimated_lines
                    max_lines = max(max_lines, total_lines)

            # Set uniform row height
            ws.row_dimensions[row_idx].height = max(30, min(max_lines * 18, 200))
    elif sheet_name == 'Agreed_disagreed_codebook':
        # For Agreed_disagreed_codebook, only calculate non-Type column content
        for row_idx in range(2, ws.max_row + 1):
            max_lines = 1
            for col_idx in range(2, ws.max_column + 1):  # Start from 2nd column
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    text = str(cell.value)
                    lines = text.count('\n') + 1
                    col_width = ws.column_dimensions[get_column_letter(col_idx)].width
                    estimated_lines = max(1, int(len(text) / (col_width)))
                    total_lines = lines + estimated_lines
                    max_lines = max(max_lines, total_lines)

            ws.row_dimensions[row_idx].height = max(30, min(max_lines * 18, 200))
    elif sheet_name == 'Decision_agreed_codebook':
        # For Decision_agreed_codebook, only calculate non-Type column content
        for row_idx in range(2, ws.max_row + 1):
            max_lines = 1
            for col_idx in range(2, ws.max_column + 1):  # Start from 2nd column
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    text = str(cell.value)
                    lines = text.count('\n') + 1
                    col_width = ws.column_dimensions[get_column_letter(col_idx)].width
                    estimated_lines = max(1, int(len(text) / (col_width)))
                    total_lines = lines + estimated_lines
                    max_lines = max(max_lines, total_lines)

            ws.row_dimensions[row_idx].height = max(30, min(max_lines * 18, 200))
    else:
        # Normal row height calculation for other sheets
        for row_idx in range(2, ws.max_row + 1):
            max_lines = 1
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    text = str(cell.value)
                    lines = text.count('\n') + 1
                    col_width = ws.column_dimensions[get_column_letter(col_idx)].width
                    estimated_lines = max(1, int(len(text) / (col_width * 1.2)))
                    total_lines = lines + estimated_lines
                    max_lines = max(max_lines, total_lines)

            ws.row_dimensions[row_idx].height = max(30, min(max_lines * 18, 300))

    # Freeze first row
    ws.freeze_panes = 'A2'

wb.save(output_file)
print(f"\n转换完成！Excel文件已保存为: {output_file}")
print(f"输入文件: {input_file}")
print(f"输出文件: {output_file}")
print(f"\n创建的工作表:")
print("1. Role_Team - 角色团队的分层代码本（角色信息为主行，代码为子行）")
print("2. Agreed_disagreed_codebook - 一致与不一致的代码")
print("3. Decision_agreed_codebook - 决策讨论结果")
print("4. Final_codebook - 最终代码本")
