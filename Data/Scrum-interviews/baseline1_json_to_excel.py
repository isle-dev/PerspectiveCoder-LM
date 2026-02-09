import json
import pandas as pd
import argparse
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Parse command line arguments
parser = argparse.ArgumentParser(description='Convert baseline JSON file to Excel format')
parser.add_argument('-i', '--input', type=str, default='baseline1.json', 
                    help='Input JSON file path')
parser.add_argument('-o', '--output', type=str, default='baseline1.xlsx', 
                    help='Output Excel file path')
args = parser.parse_args()

# Read JSON file
input_file = args.input
output_file = args.output

print(f"读取输入文件: {input_file}")
with open(input_file, 'r', encoding='utf-8') as f:
    data = json.load(f)

# Create Excel writer object
writer = pd.ExcelWriter(output_file, engine='openpyxl')

# Prepare data list
rows = []
for item in data['Codebook']:
    row = {
        'Code': item['code'],
        'Definition': item['definition'],
        'Inclusion Criteria': '\n'.join(item['inclusion_criteria']),
        'Exclusion Criteria': '\n'.join(item['exclusion_criteria']),
        'Typical Examples': '\n'.join(item['typical_examples']),
        'Atypical Examples': '\n'.join(item['atypical_examples']),
        'Participants': ', '.join(item['participants']),
        'Relevance to RQ': item['relevance_to_RQ'],
        'Notes': item['notes']
    }
    rows.append(row)

# Create DataFrame
df = pd.DataFrame(rows)
df.to_excel(writer, sheet_name='Codebook', index=False)

writer.close()

# Format Excel
wb = load_workbook(output_file)

# Define border style
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Set header row style
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Set data row format
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
            cell.border = thin_border
    
    # Auto-adjust column width
    for idx, col in enumerate(ws.columns, 1):
        column = get_column_letter(idx)
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        # Set reasonable column width
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column].width = adjusted_width
    
    # Set row height to fit content
    for row_idx in range(2, ws.max_row + 1):
        max_lines = 1
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                text = str(cell.value)
                lines = text.count('\n') + 1
                col_width = ws.column_dimensions[get_column_letter(col_idx)].width
                estimated_lines = max(1, int(len(text) / (col_width)))
                total_lines = lines + estimated_lines
                max_lines = max(max_lines, total_lines)
        
        ws.row_dimensions[row_idx].height = max(30, min(max_lines * 18, 300))
    
    # Freeze first row
    ws.freeze_panes = 'A2'

wb.save(output_file)
print(f"\n转换完成！Excel文件已保存为: {output_file}")
print(f"输入文件: {input_file}")
print(f"输出文件: {output_file}")
