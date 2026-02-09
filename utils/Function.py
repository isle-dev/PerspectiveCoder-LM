import tiktoken
import json
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
import os
import io
from pathlib import Path
from zipfile import ZipFile, ZIP_DEFLATED
import random
from config.discuss_menu import *
from config.model_menu import *

random.seed(42)


# random generate identity
def roles_identity_generate(roles_num, role=None):
    roles_identity = []
    for i in range(roles_num):
        roles_identity.append({
            "Gender": random.choice(Gender) if not role else role[i],
            "Education": random.choice(Education),
            "Race_Ethnicity": random.choice(Race_Ethnicity),
            "Age": random.choice(Age),
            "Degree_Subject": random.choice(Degree_Subject),
            "Reddit_Use": random.choice(Reddit_Use)
        })
    return roles_identity


# string tokens
def num_tokens_from_string(string: str, model_name: str) -> int:
    encoding_name = MODEL_TOKENIZER_MAP.get(model_name, "cl100k_base")
    encoding = tiktoken.get_encoding(encoding_name)
    return len(encoding.encode(string))


# open json file
def import_json(file: str):
    with open(file, "r", encoding="utf-8") as f:
        return json.load(f)


# save json file
def save_json(file: str, Target):
    with open(file, "w", encoding="utf-8") as f:
        json.dump(Target, f, indent=4)


# save all codebook to excel
def save_codebook_excel(file_path: str, target_text: str, codebook: [dict]):
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
        last_row = ws.max_row
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["target_text", "code", "definition"])
        last_row = 1

    start_row = last_row + 1
    end_row = start_row + len(codebook) - 1
    merge_range = f"A{start_row}:A{end_row}"

    # merge target_text
    ws.merge_cells(merge_range)
    ws[f"A{start_row}"] = target_text
    ws[f"A{start_row}"].alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    # load code and justification
    for idx, item in enumerate(codebook, start=start_row):
        ws.cell(row=idx, column=2, value=item["code"])  # B列：code
        ws.cell(row=idx, column=3, value=item["definition"])  # C列：justification

    # adjust column width
    ws.column_dimensions['A'].width = 80
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 80

    wb.save(file_path)
    print(f"✅ Current Codebook has been saved: {file_path}")


# save single debate process to excel
def save_discuss_excel(file_path: str, target_text: str, disagreed_list: [str], debate_list: [[str]]):
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["target_text", "disagree", "round 1", "round 2", "round 3", "round 4"])

    start_row = ws.max_row + 1

    # load debate process
    for disagree, rounds in zip(disagreed_list, debate_list):
        # 如果 rounds 不足 4，则补空；如果多于 4，截断
        while len(rounds) < 4:
            rounds.append("")
        if len(rounds) > 4:
            rounds = rounds[:4]
        row = [target_text, str(disagree)] + rounds
        ws.append(row)

    end_row = ws.max_row

    if end_row > start_row:
        merge_range = f"A{start_row}:A{end_row}"
        ws.merge_cells(merge_range)
        ws[f"A{start_row}"].alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    else:
        ws[f"A{start_row}"].alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    # adjust column width
    ws.column_dimensions['A'].width = 80
    ws.column_dimensions['B'].width = 40
    for col in ['C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 80

    wb.save(file_path)
    print(f"✅ Excel with merged target_text saved: {file_path}")


def zip_folder_to_bytes(folder_path: str) -> bytes:
    folder = Path(folder_path)
    buf = io.BytesIO()
    with ZipFile(buf, "w", ZIP_DEFLATED) as zf:
        for p in folder.rglob("*"):
            if p.is_file():
                zf.write(p, p.relative_to(folder).as_posix())
    buf.seek(0)
    return buf.read()

